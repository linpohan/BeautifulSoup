# -*- coding: utf-8 -*-
"""
Created on Sun Sep 19 13:05:15 2021

@author: User
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from openpyxl.drawing.image import Image

wb = Workbook()
ws = wb.active

url = 'https://supertaste.tvbs.com.tw/food#'
data = requests.get(url, 'html.parser')
data.encoding = 'UTF-8'
data = data.text
sp = BeautifulSoup(data)

foods = sp.find(id='combolistUl').find_all('li')

time = []
title = []
web = []

for row in foods:
    try:
        time.append(row.find('div', class_ = 'time').text.strip())
        title.append(row.find('div', class_ = 'txt').text)
        web.append('https://supertaste.tvbs.com.tw' + row.find('a').get('href'))
    except Exception as e:
        print(e)


highway = pd.DataFrame({
                        "時間":time,
                        "標題":title,
                        "網址":web
    }, columns=[
                "圖示","時間","標題","網址"
        ])
highway.to_excel('Tvbs-food.xlsx', index=False)

wb = load_workbook('Tvbs-food.xlsx')
ws = wb.worksheets[0]
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 94
ws.column_dimensions['D'].width = 44


col = 2
for row in foods:
    try:
        imgs = row.find('img').get('data-original')
        imgUrl = requests.get(imgs)
        fileName = imgs.split('/')[-1]
        if not os.path.exists('food-images'):
            os.mkdir('food-images')
        with open('food-images\\' + fileName, 'wb') as f:
            f.write(imgUrl.content)
        
        imageSave = Image('food-images\\' + fileName)
        imageSave.width, imageSave.height = 164,164
        ws.column_dimensions['A'].width = 24
        ws.row_dimensions[col].height = 130
        ws.add_image(imageSave, 'A{}'.format(col))
        col += 1
    except Exception as e:
        print(e)


wb.save('Tvbs-food.xlsx')