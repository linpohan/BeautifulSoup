# -*- coding: utf-8 -*-
"""
Created on Sat Sep 18 19:42:53 2021

@author: User
"""

import requests
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pandas as pd


wb = Workbook()
ws = wb.active

url = 'https://tw.mall.yahoo.com/APPLE%E8%98%8B%E6%9E%9C-%E6%89%8B%E6%A9%9F-%E6%99%BA%E6%85%A7%E7%A9%BF%E6%88%B4-794017232-category.html?.r=1207549978'
data = requests.get(url, 'html.parser')
data.encoding = 'UTF-8'
data = data.text
sp = BeautifulSoup(data)

allPhone = sp.find(id='Mid_item_list')
phones = allPhone.find_all('div', class_ = 'Bfc Pt-16 Bd-end')

title = []
price = []
web = []

for row in phones:
    title.append(row.find('a').text)
    price.append(row.find('span').text)
    web.append(row.find('a').get('href'))
    
highway = pd.DataFrame({
                        "標題":title,
                        "價位":price,
                        "網址":web
    },columns=[
                "圖片",
                "標題",
                "價位",
                "網址"
        ]
    )
        

highway.to_excel("yahooshop-apple.xlsx", encoding="utf-8", index=False)

wb = load_workbook('yahooshop-apple.xlsx')
ws = wb.worksheets[0]
ws.column_dimensions['B'].width = 90
ws.column_dimensions['D'].width = 50

imgs = allPhone.find_all('div', class_ = 'Fl-start Pos-r')
col = 2
for row in imgs:
    img = row.find('img').get('src')
    imgUrl = requests.get(img)
    fileName = img.split('/')[-1]
    
    if not os.path.exists('images-apple'):
        os.mkdir('images-apple')
    with open('images-apple\\' + fileName, 'wb') as f:
        f.write(imgUrl.content)
        
    imgSave = Image('images-apple\\' + fileName)
    imgSave.width, imgSave.height = 164, 164
    ws.column_dimensions['A'].width = 24
    ws.row_dimensions[col].height = 130
    ws.add_image(imgSave, 'A{}'.format(col))
    col = col + 1
    
wb.save('yahooshop-apple.xlsx')


